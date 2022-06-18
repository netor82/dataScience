#!/usr/bin/env python
# coding: utf-8

# # Proyecto
# 
# **Big data**
# 
# Profesor: Luis Alexander Calvo Valverde
# 
# Instituto Tecnológico de Costa Rica
# 
# Estudiante: Ernesto Rivera
# 

# # Datos elegidos
# 
# ## SAFIs: Utilidades y carteras financieras
# Las Sociedades Administradoras de Fondos de Invesión **SAFIs**, son supervisadas por la Superintendencia General de Valores **SUGEVAL**.
# 
# Cada entidad supervisada debe presentar **mensualmente** informes sobre la composición de su cartera así como información de balances financieros (ficha financiera).
# 
# Esta información es compilada y publicada por la SUGEVAL en sus páginas sobre [Carteras](https://aplicaciones.sugeval.fi.cr/InformesEstadisticas/CarterasSAFI) y [Fichas Financieras](https://aplicaciones.sugeval.fi.cr/InformesEstadisticas/SituacionFinancieraSAFI).
# 
# La información disponible puede ser exportada _manualmente_ en formato xlsx (Microsoft Excel) en dichas páginas.
# 
# ## Archivos de entrada
# Los archivos están en formato xlsx.  En todos los archivos, que se encuentran en la carpeta `/src/data` dentro del contenedor o dentro de los archivos incluidos en el zip.
# 
# ### 1. Fichas Financieras
# Son un par de archivos que contienen la información financiera de las SAFIs:
# - FichaFinanciera_1 desde 01/2017 hasta 12/2019
# - FichaFinanciera_2 desde 01/2020 hasta 03/2022
# 
# ### 2. Composición de la Cartera
# La información disponible inicia en 01/2018 hasta 03/2022.  Existe un archivo por cada mes de información en dicho periodo de tiempo.  El nombre de los archivos sigue el formato `CarteraEntidades_YYYYMM.xlsx`, siendo YYYY el año y MM el mes.
# 
# ### Procesamiento
# Se utiliza la biblioteca de python `openpyxl` para extraer la información de los archivos.  En la sección "Lectura y Limpieza" se detalla las columnas y filas que son utilizadas.
# 
# ### Importar bibliotecas de python

# In[1]:


from datetime import datetime
from decimal import Decimal, ConversionSyntax
import glob

import findspark
#-#findspark.init()

import pyspark

from pyspark.sql import SparkSession
from pyspark.sql.types import (StringType, IntegerType, FloatType, DateType,
                               DecimalType, StructField, StructType)
from pyspark.sql.functions import avg, when, col, isnan, count, add_months, unix_timestamp
from pyspark.ml.feature import StandardScaler, StringIndexer, VectorAssembler
from pyspark.ml.classification import LogisticRegression
from pyspark.ml.evaluation import RegressionEvaluator
from pyspark.ml.regression import DecisionTreeRegressor

from pyspark.ml.stat import Correlation
from pyspark.ml.functions import vector_to_array
import seaborn as sns
import matplotlib.pyplot as plt

import openpyxl


# ## Para cargar los datos en diferentes sistemas operativos
# 
# Comente o descomente las siguientes líneas dependiendo en cuál sistema operativo tiene docker corriendo.

# In[2]:


# mac o windows
#-#database_url = "jdbc:postgresql://host.docker.internal:5433/postgres"

# linux
#-#database_url = "jdbc:postgresql://172.17.0.1:5433/postgres"


# ____
# 
# # Lectura y limpieza
# 
# ____
# 
# ## 1. Leer datos de Excel
# 
# Iniciamos con las funciones para leer los datos de los archivos de Excel.
# 
# De los archivos de **Ficha Financiera**, solo vamos a leer:
#  - El código de la SAFI, que está en el nombre de la hoja de excel.
#  - El margen de utilidad neta, está en la fila 16.
#  - El mes al que cada margen corresponde.  Los meses están por columnas, en la fila 12.
#  
# De los archivos de **Cartera de las Entidades**, se van a leer las siguientes columnas:
#  - A=0: Código de cada SAFI.
#  - C=2: Fecha del reporte.
#  - I=8: Tipo de título:
#     1. Renta fija (deuda)
#     2. Renta Variable (acciones)
#     3. Valores de participación fondos cerrados
#     4. Valores de participación fondos abiertos
#     5. Fideicomiso de Titularización de deuda
#     6. Fideicomiso de Titularización de participación
#     7. Fideicomiso de Titularización valores mixtos
#     8. Productos Estructurados
#  - O=14: Sector Emisor:
#     1. Público
#     2. Privado
#  - P=15: Moneda, como USD o CRC.
#  - U=20: Premio, puntos adicionales que se suman a la tasa de referencia en instrumentos de deuda cuya tasa de cupón está indexada
#  - W=22: Costo, desembolso o pago por la adquisición del título. Está en la moneda de la columna P.
#  - Y=24: Ganancias Pérdidas, monto de las ganancias o pérdidas no realizadas en valoración a precios de mercado, a la fecha de corte  de la cartera.
#  - Z=25: Monto Deterioro.
#  - AA=26: Valor de Mercado, precio de valoración de un título, con base en el precio observado de las transacciones de mercado.
#  - AB=27: Valor de Mercado en colones. Columna AA convertida a colones usando el tipo de cambio de referencia para la fecha del informe.
# 
# La definición de estas columnas se encuentran más detalladas en el [archivo de glosario](https://www.sugeval.fi.cr/informesmercado/Documents/glosario.docx) provisto por SUGEVAL
# 
# Las columnas no incluidas en la lista anterior se descartan por diferentes razones:
#  - Nombre de la entidad está directamente relacionada con columna A, el código único de la SAFI.
#  - Fechas del reporte o de compra de la inversión.
#  - Mercado secundario o primario.  Dónde se compró o quién el instrumento es irrelevante.  Lo mismo aplica para el ID de la operación, código y nombre del Emisor.
#  - Nombre del instrumento y código, tampoco se necesita para predecir utilidades.
#  - Valor Facial: la definición de la columna dice que corresponde al valor nominal, sin embargo tenemos el valor del mercado (AA=26). El valor facial se registra cuando se emitió el valor. Como no cuenta depreciación es mejor ignorarlo.
#  - Otras columnas solo aplican para instrumentos de tipo deuda, donde no hay datos o son datos por defecto para otros tipos de instrumentos.
#  - Por último, algunas columnas son acumulados de intereses.

# In[3]:


def leerFichasFinancieras():
    lista = []
    leerFichaFinanciera('./data/FichaFinanciera_1.xlsx', lista)
    leerFichaFinanciera('./data/FichaFinanciera_2.xlsx', lista)
    return lista

def leerFichaFinanciera(archivo, lista):
    wb = openpyxl.load_workbook(archivo)

    for i in range(3, len(wb.sheetnames)):
        sheetName = wb.sheetnames[i]
        sheet = wb[sheetName]
        
        # inicia en columna 'C', e itera por todas las columnas.
        columnaActual = 3
        # filas fija
        fecha = sheet.cell(12, columnaActual).value
        
        while fecha:
            margenUtilidad = sheet.cell(16, columnaActual).value
            
            try:
                if margenUtilidad:
                    lista.append((sheetName,
                            datetime.strptime(fecha.strip(), '%Y/%m'),
                            float(margenUtilidad)))
            except BaseException as valueError:
                print('Error en', archivo, sheetName, columnaActual, margenUtilidad, valueError)
            
            columnaActual += 1
            fecha = sheet.cell(12, columnaActual).value
        
    wb.close()
    return lista


def leerCartera(archivo):
    wb = openpyxl.load_workbook(archivo)
    sheet = wb.active
    
    lista = []
    
    # guarda solo algunas columnas de interés
    columnas = [0, 2, 8, 14, 15, 20, 22, 24, 25, 26, 27]
    for fila in sheet.iter_rows(min_row=12):
        valoresFila = []
        for columna in columnas:
            if columna == 2: # fecha
                # todos queden en el 1o del mes
                valor = fila[columna].value.replace(day=1)
            elif 20 <= columna: #columnas de monedas
                valor = Decimal(fila[columna].value)
            else:
                valor = fila[columna].value
            valoresFila.append(valor)

        lista.append(tuple(valoresFila))
    wb.close()
    return lista


# ## 2. Crear sesión de Spark y carga de datos
# El resultado final debe ser un par de DataFrames de Spark con toda la información leída de los archivos.

# In[4]:


spark = SparkSession \
   .builder \
   .appName("Basic JDBC pipeline") \
   .config("spark.driver.extraClassPath", "postgresql-42.2.14.jar") \
   .config("spark.executor.extraClassPath", "postgresql-42.2.14.jar") \
   .getOrCreate()


# In[5]:


def crearDfFichaFinanciera(infoFichaFinanciera):
    schema = StructType([
                StructField("cod", StringType()),
                StructField("mes", DateType()),
                StructField("utilidad", FloatType())
        ])
    return spark.createDataFrame(data=infoFichaFinanciera, schema=schema)

def crearDfCartera(infoCartera):
    schema = StructType([
                StructField("cod", StringType()),
                StructField("mes", DateType()),
                StructField("tipoTitulo", IntegerType()),
                StructField("sector", IntegerType()),
                StructField("moneda", StringType()),
                StructField("premio", DecimalType(12, 2)),
                StructField("costo", DecimalType(12, 2)),
                StructField("ganacias", DecimalType(12, 2)),
                StructField("montoDeterioro", DecimalType(12, 2)),
                StructField("valorMercado", DecimalType(12, 2)),
                StructField("valorMercadoCol", DecimalType(12, 2))
        ])
    return spark.createDataFrame(data=infoCartera, schema=schema)


def leerCarteras():
    # DF vacío
    cartera_df = crearDfCartera([])
    
    for archivo in glob.glob("/src/data/CarteraEntidades*.xlsx"):
        cartera = leerCartera(archivo)
        unaCartera_df = crearDfCartera(cartera)
        cartera_df = cartera_df.union(unaCartera_df)
    
    return cartera_df

#-#margenUtilidad_df = crearDfFichaFinanciera(leerFichasFinancieras())
#-#margenUtilidad_df.printSchema()

#-#cartera_df = leerCarteras()
#-#cartera_df.printSchema()


# ## 3. Limpieza de datos
# 
# Comprobaciones realizadas:
# - Todas las entidades están reportadas en ambos datasets.
# - Todas las entidades recabadas en el DF de **Fichas Financieras** contienen 63 meses de información de utilidades.
# - No existen valores nulos en los datasets.
# - Se analizó las columnas _montoDeterioro_ y _premio_ por tener muchos ceros. Pero decidí no quitarlas antes de analizar resultados.
# 
# Existe una SAFI _GLSAFI_ que aparece con únicamente indicador de liquidez para los últimos dos meses.  Sin embargo, el proceso de lectura lo descartó por no tener datos de utilidades.  Esta entidad tampoco aparece en los archivos de composición de cartera.
# 
# Limpiezas realizadas:
# - El reporte de utilidad utiliza el primer día del mes, el reporte de cartera el último. Se elimina el día del mes del de cartera.  Este proceso se hizo durante la lectura de los archivos excel.

# In[6]:


#-#margenUtilidad_df.show(5)
#-#cartera_df.show(5)


# In[7]:


# ambos datasets tienen los mismos códigos de SAFIs
#-#"Son iguales" if cartera_df.select('cod').distinct().orderBy('cod').collect() == \
#    margenUtilidad_df.select('cod').distinct().orderBy('cod').collect() else "uuups"


# In[8]:


# todos los SAFIs tienen exactamente la misma cantidad de información
# (los mismos meses de información)
#-#margenUtilidad_df.groupBy('cod').agg(count('utilidad').alias('count')).orderBy('cod').show()


# In[9]:


def contarNulos(df):
    df2 = df.select([count(when(col(c).isNull(), c)).alias(c) for c in df.columns])
    df2.show()
    return df2

#-#contarNulos(margenUtilidad_df)
#-#contarNulos(cartera_df)


# In[10]:


# diferentes valores en montoDeterioro
#-#cartera_df.groupBy('montoDeterioro').agg(count('montoDeterioro').alias('count')).orderBy(col('count').desc()).show(15)


# In[11]:


# diferentes valores en 'premio'
#-#cartera_df.groupBy('premio').agg(count('premio').alias('count')).orderBy(col('count').desc()).show(15)


# In[12]:


#-#cartera_df = cartera_df.select('cod','mes','tipoTitulo','sector',
#                              'moneda','premio',
#                              'costo','ganacias','montoDeterioro',
#                              'valorMercado','valorMercadoCol')


# ## 4. Features textual a códigos numéricos
# Esta transformación solo aplica para el dataframe de Cartera de Inversiones.

# In[13]:


def deTextACodigos(df):
    label_stringIdx = StringIndexer(inputCols = ['cod', 'moneda'], outputCols = ['cod2', 'moneda_2'])
    model = label_stringIdx.fit(df)
    indexadas = model.transform(df)
    return indexadas

#-#cartera_df = deTextACodigos(cartera_df)
#-#cartera_df.select(['cod', 'cod2','moneda', 'moneda_2']).show(10)


# ## 5. Todos los montos monetarios convertidos a colones
# 
# El tipo de cambio se pone en una nueva columna.  Con él, se calculan el resto de campos a colones.
# 
# Esta transformación solo aplica para el dataframe de Cartera de Inversiones.

# In[14]:


def calcularColones(df):
    conColones = df \
        .withColumn('tipoCambio', (col('valorMercadoCol') / col('valorMercado')).cast(DecimalType(12,2))) \
        .withColumn('premioCol', (col('premio') * col('tipoCambio')).cast(DecimalType(12,2))) \
        .withColumn('costoCol', (col('costo') * col('tipoCambio')).cast(DecimalType(12,2))) \
        .withColumn('ganaciasCol', (col('ganacias') * col('tipoCambio')).cast(DecimalType(12,2))) \
        .withColumn('montoDeterioroCol', (col('montoDeterioro') * col('tipoCambio')).cast(DecimalType(12,2)))
    return conColones

#-#cartera_tipoDeCambio_df = calcularColones(cartera_df)
#-#cartera_tipoDeCambio_df.select(['tipoCambio', 'moneda', 'premio','premioCol', 'costo', 'costoCol']).show(10)
#-#cartera_tipoDeCambio_df.printSchema()


# ## 6. Pegar t con t+1 en preparación para analizar como serie de tiempo
# Como resultado de este proceso, la cantidad de registros disminuye en 14 pues son 14 las SAFIs.
# 
# La columna 'label' es el _target_ a predecir

# In[15]:


def margenUtilidadSiguienteMes(df):
    df1 = df.select(['cod','mes',add_months(col('mes'), 1).alias('sig_mes'),'utilidad']).alias('a')
    df2 = df.alias('b')
    result = df2 \
        .join(df1, [col('a.cod') == col('b.cod'), col('a.sig_mes') == col('b.mes')], 'inner') \
        .select('a.cod', 'a.mes', 'a.sig_mes', 'a.utilidad', col('b.utilidad').alias('label'))
    return result

#-#print('Previo:')
#-#print('cantidad inicial', margenUtilidad_df.count())
#-#margenUtilidad_df.show(5)
#-#print()

#-#print('Después:')
#-#margenUtilidadSiguiente_df = margenUtilidadSiguienteMes(margenUtilidad_df)
#-#margenUtilidadSiguiente_df.filter(col('cod') == 'ALDESASFI').orderBy(col('mes').asc()).show(4)
#-#print('cantidad final', margenUtilidadSiguiente_df.count())


# ## 7. Pegar ambos datasets

# In[16]:


def unirParaAnalizar(cartera_df, utilidades_df):
    dfc = cartera_df.alias('c')
    dfu = utilidades_df.alias('u')
    
    return dfc \
        .join(dfu, [col('c.cod') == col('u.cod'), col('c.mes') == col('u.mes')], 'inner') \
        .select('c.cod', 'c.mes', 'c.cod2', 'tipoTitulo', 'sector', 'moneda_2', 
                'valorMercadoCol', 'premioCol', 'costoCol',
                'ganaciasCol', 'montoDeterioroCol',
                'u.utilidad', 'u.label')

#-#joint_df = unirParaAnalizar(cartera_tipoDeCambio_df, margenUtilidadSiguiente_df)
#-#joint_df.show(8)


# ____
# 
# # Guardar en Postgres
# 
# ____
# 
# ## 1. Gardar los datos
# 

# In[17]:


def salvarDb(df, nombreTabla):
    df\
    .write \
    .format("jdbc") \
    .mode("overwrite") \
    .option("url", database_url) \
    .option("user", "postgres") \
    .option("password", "testPassword") \
    .option("dbtable", nombreTabla) \
    .save()

#-#salvarDb(joint_df, 'infoSafi')


# ## 2. Leer los datos

# In[18]:


def leerDb(nombreTabla):
    df = spark \
    .read \
    .format("jdbc") \
    .option("url", database_url) \
    .option("user", "postgres") \
    .option("password", "testPassword") \
    .option("dbtable", nombreTabla) \
    .load()
    
    return df

#-#infoSafi = leerDb('infoSafi')
#-#infoSafi.show(2)


# ____
# 
# # Modelos de Predicción
# 
# ____
# 
# ## 1. Preparación: Vectorizar los datos
# 

# In[19]:


# no se pueden vectorizar fechas, entonces se convierten en Epocs
#-#infoSafi2 = infoSafi.withColumn('mes2', unix_timestamp(col('mes')))

# se usa la columna 'mes2' en vez de 'mes'
#-#assembler = VectorAssembler(
#    inputCols=['cod2','mes2','tipoTitulo','sector','moneda_2','valorMercadoCol','premioCol',
#               'costoCol','ganaciasCol','montoDeterioroCol','utilidad','label'],
#    outputCol="features")

#-#vector_df = assembler.transform(infoSafi2)

#-#vector_df.show(5)


# ## 2. Preparación: Estandarización de los datos

# In[20]:


#-#standard_scaler = StandardScaler(inputCol='features', outputCol='scaled')
#-#scale_model = standard_scaler.fit(vector_df)
#-#scaled_df = scale_model.transform(vector_df)
#-#scaled_df.show(5)


# In[21]:


#-#scaled_df.show(5)


# ## 3. Preparación: Separar datos de entrenamiento y prueba
# 
# Por ser una serie de tiempo, se toman los últimos meses como datos de pruebas y el resto como datos de entrenamiento.

# In[22]:


#(trainingData, testData) = scaled_df.randomSplit([0.6, 0.4], 1234)

#-#fechaLimite = datetime(2022, 2, 1)
#-#trainingData = scaled_df.filter(col('mes') < fechaLimite)
#-#testData = scaled_df.filter(col('mes') >= fechaLimite)

#-#print("Elementos para entrenamiento", trainingData.count())
#-#print("Elementos para pruebas", testData.count())

# valores por defecto: labelCol="label", predictionCol="prediction", metricName="rmse"
#-#evaluator = RegressionEvaluator()


# In[23]:


def regresionYAnalisis(rg):
    model = rg.fit(trainingData)
    predictions = model.transform(testData)
    rmse = evaluator.evaluate(predictions)
    print("Root Mean Squared Error (RMSE) on test data = %g" % rmse)
    predictions.select('cod','mes','label', 'prediction').show(40)
    
    # ordenar predicciones para tomar la mejor SAFI
    print("Predicción promedio por SAFI a partir de los datos de prueba - último mes:")
    predictions.select('cod','prediction').groupBy('cod').agg(avg('prediction').alias('p')).orderBy(col('p').desc()).show(5)


# ## 4. Regresión Lineal 

# In[24]:


from pyspark.ml.regression import LinearRegression
def regresionLineal():
    lr = LinearRegression(maxIter=10, regParam=0.3, elasticNetParam=0.8)
    lrModel = lr.fit(scaled_df)
    trainingSummary = lrModel.summary
    print("Root Mean Squared Error (RMSE): %f" % trainingSummary.rootMeanSquaredError)

#-#regresionLineal()


# 
# ## 5. Regresión con Árbol de Decisión

# In[25]:


def regresionConArbolDeDecision():
    tree_reg = DecisionTreeRegressor(featuresCol="scaled")
    regresionYAnalisis(tree_reg)

#-#regresionConArbolDeDecision()


# ## 6. Regresión con Random Forest

# In[26]:


from pyspark.ml.regression import RandomForestRegressor

def regresionConRandomForest():
    rf = RandomForestRegressor(featuresCol="scaled")
    regresionYAnalisis(rf)

#-#regresionConRandomForest()


# # Análisis de resultados
# 
# ## Sobre el proceso
# 
# La pricipal dificultad fue entender los datos y los términos del dominio del negocio en cuestión.  Decisiones como cuáles features no son necesarios requirieron investigación.
# 
# El trabajo con series de tiempo es un poco distinto que las regresiones regulares:
# - El orden importa.
# - La división de datos de entrenamiento y pruebas debe respetar el orden temporal.
# - Queda fuera de este trabajo, pero el modelo debe actualizarse cada vez que se reciban datos nuevos.
# 
# ## Sobre las fuentes de datos
# 
# - Los datos provistos por la SUGEVAL son de buena calidad entanto no contienen datos y están bien registrados.
# - Sin embargo, por el formato de archivo, se necesitó código para extraer la información de los archivos.  En este mismo proceso se hizo tranformaciones como cambios de tipos de datos o hacer que todas las fechas estuvieran en el primer día del mes en lugar del último.
# - Recomendada la biblioteca openpyxl para leer datos, es muy intuitiva.
# 
# ## Sobre los resultados
# 
# Para ambos métodos de regresión utilizados (Decision Tree, Random Tree Forest), concuerdan con los datos del último mes de utilidades: la SAFI con código 'INSBANCREDITSFI', o INS-Inversiones SAFI SA, es la que se proyecta con los mejores rendimientos.
# 
# A continuación el top 5 de mejores utilidades con los datos de entrada para marzo (leído de los datos de SUGEVAL sin procesar).

# In[27]:


#-#margenUtilidad_df.filter(col('mes') > fechaLimite).orderBy(col('utilidad').desc()).show(5)

